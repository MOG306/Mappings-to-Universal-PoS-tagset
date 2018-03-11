
import openpyxl

def getList(s):
    L = map(str, s.split())
    return L

def getPrinted(e):
    ans = ""
    for item in e:
        if (len(ans) > 0):
            ans += " - "
        ans += item
    return ans

book_conv = openpyxl.load_workbook('CONVERSION.xlsx')
book_code = openpyxl.load_workbook('CODE.xlsx')
book_out = openpyxl.Workbook()


sheet_out = book_out.active
sheet_conv = book_conv.active
sheet_code = book_code.active

code_mapping = {}
for r in range(2, sheet_code.max_row + 1):
    x = sheet_code.cell(row = r, column = 1).value
    y = sheet_code.cell(row = r, column = 2).value
    if (x in code_mapping):
        code_mapping[x][y] = 1
    else:
        code_mapping[x] = {y : 1}
for item in code_mapping:
    code_mapping[item] = getPrinted(code_mapping[item])

for r in range(2, sheet_conv.max_row + 1):
    L = getList(sheet_conv.cell(row = r, column = 2).value)
    e = {}
    for item in L:
        if (item in code_mapping):
            e[code_mapping[item]] = 1
    sheet_conv['C' + str(r)] = getPrinted(e)


book_conv.save("OUTPUT.xlsx")
